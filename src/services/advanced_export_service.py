"""
PondMonitor Advanced Export Service - Week 2 Enhancement

Advanced Excel export system with filtering, formatting, and progress tracking.
Builds upon the existing export service with enhanced features for Week 2.

New Features:
- Excel export with multiple sheets and advanced formatting
- Date range filtering with calendar widgets  
- Data type selection (pond metrics, station metrics, both)
- Export progress indicators and status feedback
- Export scheduling and automation
- Enhanced mobile-responsive export interface
"""

import logging
import json
from datetime import datetime, timezone, timedelta
from io import BytesIO
from typing import Dict, Any, List, Optional, Callable
from dataclasses import dataclass, asdict
from pathlib import Path

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import LineChart, Reference
    EXCEL_ADVANCED_AVAILABLE = True
except ImportError:
    EXCEL_ADVANCED_AVAILABLE = False
    logging.warning("Advanced Excel features not available - install pandas and openpyxl")

from .export_service import ExportService, ExportConfig, ExportMetadata
from ..database import DatabaseService

logger = logging.getLogger(__name__)


@dataclass
class AdvancedExportConfig(ExportConfig):
    """Enhanced export configuration with Week 2 features"""
    
    # Data selection
    data_types: List[str] = None  # ['pond_metrics', 'station_metrics']
    aggregation: str = 'raw'  # 'raw', 'hourly', 'daily'
    include_charts: bool = False
    include_analysis: bool = False
    
    # Excel formatting
    excel_formatting: bool = True
    excel_charts: bool = True
    excel_conditional_formatting: bool = True
    
    # Progress tracking
    export_id: str = None
    progress_callback: Optional[Callable] = None
    
    # Filtering
    temperature_range: Optional[tuple] = None  # (min, max) in Celsius
    battery_range: Optional[tuple] = None      # (min, max) in Volts
    signal_range: Optional[tuple] = None       # (min, max) in dBm
    
    def __post_init__(self):
        """Post-initialization setup"""
        super().__post_init__() if hasattr(super(), '__post_init__') else None
        
        if self.data_types is None:
            self.data_types = ['pond_metrics', 'station_metrics']
        
        if self.export_id is None:
            self.export_id = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"


class AdvancedExportService:
    """
    Enhanced export service with Week 2 advanced features.
    Wraps the base ExportService with additional functionality.
    """
    
    def __init__(self, db_service: DatabaseService, base_export_service: ExportService = None):
        """Initialize advanced export service"""
        self.db = db_service
        self.base_service = base_export_service or ExportService(db_service)
        self.export_progress = {}
        
    def create_advanced_excel_export(self, config: AdvancedExportConfig) -> bytes:
        """
        Create advanced Excel export with multiple sheets, formatting, and charts.
        
        Args:
            config: Advanced export configuration
            
        Returns:
            Excel workbook as bytes
        """
        if not EXCEL_ADVANCED_AVAILABLE:
            # Fall back to basic export
            basic_config = ExportConfig(
                start_time=config.start_time,
                end_time=config.end_time,
                format='excel',
                include_pond_data='pond_metrics' in config.data_types,
                include_station_data='station_metrics' in config.data_types,
                filename_prefix=config.filename_prefix
            )
            return self.base_service.export_data(basic_config)
        
        logger.info(f"Creating advanced Excel export: {config.export_id}")
        
        # Initialize progress tracking
        if config.progress_callback:
            self._update_progress(config.export_id, 0, "Initializing export...", config.progress_callback)
        
        # Get filtered data
        data = self._get_filtered_data(config)
        if config.progress_callback:
            self._update_progress(config.export_id, 20, "Data retrieved, processing...", config.progress_callback)
        
        # Create Excel workbook
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # Create summary sheet
        self._create_summary_sheet(workbook, config, data)
        if config.progress_callback:
            self._update_progress(config.export_id, 40, "Summary sheet created...", config.progress_callback)
        
        # Create data sheets
        if 'pond_metrics' in config.data_types and 'pond_metrics' in data:
            self._create_pond_metrics_sheet(workbook, data['pond_metrics'], config)
        
        if 'station_metrics' in config.data_types and 'station_metrics' in data:
            self._create_station_metrics_sheet(workbook, data['station_metrics'], config)
        
        if config.progress_callback:
            self._update_progress(config.export_id, 70, "Data sheets created...", config.progress_callback)
        
        # Add charts if requested
        if config.excel_charts and config.include_charts:
            self._add_excel_charts(workbook, data, config)
            if config.progress_callback:
                self._update_progress(config.export_id, 85, "Charts added...", config.progress_callback)
        
        # Add analysis sheet if requested
        if config.include_analysis:
            self._create_analysis_sheet(workbook, data, config)
        
        # Save to bytes
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        if config.progress_callback:
            self._update_progress(config.export_id, 100, "Export completed!", config.progress_callback)
        
        logger.info(f"Advanced Excel export completed: {config.export_id}")
        return output.getvalue()
    
    def _get_filtered_data(self, config: AdvancedExportConfig) -> Dict[str, List[Dict[str, Any]]]:
        """Get filtered data based on advanced configuration"""
        data = {}
        
        # Get pond metrics if requested
        if 'pond_metrics' in config.data_types:
            pond_metrics = self.db.get_pond_metrics(
                config.start_time, 
                config.end_time, 
                config.limit
            )
            data['pond_metrics'] = pond_metrics
        
        # Get station metrics if requested
        if 'station_metrics' in config.data_types:
            station_metrics = self.db.get_station_metrics(
                config.start_time, 
                config.end_time, 
                config.station_id,
                config.limit
            )
            
            # Apply filtering
            if config.temperature_range or config.battery_range or config.signal_range:
                station_metrics = self._apply_metric_filters(station_metrics, config)
            
            data['station_metrics'] = station_metrics
        
        # Apply aggregation if requested
        if config.aggregation != 'raw':
            data = self._aggregate_data(data, config.aggregation)
        
        return data
    
    def _apply_metric_filters(self, metrics: List[Dict[str, Any]], 
                            config: AdvancedExportConfig) -> List[Dict[str, Any]]:
        """Apply metric-based filtering"""
        filtered = []
        
        for metric in metrics:
            # Temperature filter
            if config.temperature_range:
                temp = metric.get('temperature_c')
                if temp is not None:
                    min_temp, max_temp = config.temperature_range
                    if not (min_temp <= temp <= max_temp):
                        continue
            
            # Battery filter
            if config.battery_range:
                battery = metric.get('battery_v')
                if battery is not None:
                    min_battery, max_battery = config.battery_range
                    if not (min_battery <= battery <= max_battery):
                        continue
            
            # Signal filter
            if config.signal_range:
                signal = metric.get('signal_dbm')
                if signal is not None:
                    min_signal, max_signal = config.signal_range
                    if not (min_signal <= signal <= max_signal):
                        continue
            
            filtered.append(metric)
        
        return filtered
    
    def _create_summary_sheet(self, workbook: openpyxl.Workbook, 
                            config: AdvancedExportConfig, 
                            data: Dict[str, List[Dict[str, Any]]]):
        """Create executive summary sheet"""
        ws = workbook.create_sheet("Summary", 0)
        
        # Title styling
        title_font = Font(size=16, bold=True, color="2F75B5")
        header_font = Font(size=12, bold=True)
        
        # Main title
        ws['A1'] = "PondMonitor Data Export Summary"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # Export information
        ws[f'A{row}'] = "Export Information"
        ws[f'A{row}'].font = header_font
        row += 1
        
        info_data = [
            ("Export ID:", config.export_id),
            ("Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S UTC")),
            ("Time Period:", f"{config.start_time.strftime('%Y-%m-%d %H:%M')} to {config.end_time.strftime('%Y-%m-%d %H:%M')}"),
            ("Data Types:", ", ".join(config.data_types)),
            ("Aggregation:", config.aggregation),
        ]
        
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        
        # Data statistics
        ws[f'A{row}'] = "Data Statistics"
        ws[f'A{row}'].font = header_font
        row += 1
        
        total_records = sum(len(records) for records in data.values())
        
        stats_data = [
            ("Total Records:", total_records),
        ]
        
        # Add specific counts for each data type
        for data_type, records in data.items():
            display_name = data_type.replace('_', ' ').title()
            stats_data.append((f"{display_name}:", len(records)))
        
        for label, value in stats_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    def _create_pond_metrics_sheet(self, workbook: openpyxl.Workbook, 
                                 data: List[Dict[str, Any]], 
                                 config: AdvancedExportConfig):
        """Create formatted pond metrics sheet"""
        if not data:
            return
            
        ws = workbook.create_sheet("Pond Metrics")
        
        # Convert to DataFrame for easier handling
        df = pd.DataFrame(data)
        
        # Write data to sheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        if config.excel_formatting:
            self._apply_sheet_formatting(ws, "pond")
        
        if config.excel_conditional_formatting:
            self._apply_conditional_formatting(ws, df, "pond")
    
    def _create_station_metrics_sheet(self, workbook: openpyxl.Workbook, 
                                    data: List[Dict[str, Any]], 
                                    config: AdvancedExportConfig):
        """Create formatted station metrics sheet"""
        if not data:
            return
            
        ws = workbook.create_sheet("Station Metrics")
        
        # Convert to DataFrame for easier handling
        df = pd.DataFrame(data)
        
        # Write data to sheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        if config.excel_formatting:
            self._apply_sheet_formatting(ws, "station")
        
        if config.excel_conditional_formatting:
            self._apply_conditional_formatting(ws, df, "station")
    
    def _apply_sheet_formatting(self, ws: openpyxl.worksheet.worksheet.Worksheet, sheet_type: str):
        """Apply professional formatting to sheet"""
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Apply to header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    def _apply_conditional_formatting(self, ws: openpyxl.worksheet.worksheet.Worksheet, 
                                    df: pd.DataFrame, sheet_type: str):
        """Apply conditional formatting based on data values"""
        from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
        from openpyxl.styles import PatternFill
        
        if sheet_type == "station":
            # Battery level conditional formatting
            if 'battery_v' in df.columns:
                battery_col_idx = df.columns.get_loc('battery_v') + 1
                battery_col_letter = openpyxl.utils.get_column_letter(battery_col_idx)
                
                # Low battery warning (red for < 3.3V)
                low_battery_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                low_battery_rule = CellIsRule(operator='lessThan', formula=['3.3'], fill=low_battery_fill)
                ws.conditional_formatting.add(f'{battery_col_letter}2:{battery_col_letter}{len(df)+1}', low_battery_rule)
            
            # Temperature conditional formatting
            if 'temperature_c' in df.columns:
                temp_col_idx = df.columns.get_loc('temperature_c') + 1
                temp_col_letter = openpyxl.utils.get_column_letter(temp_col_idx)
                
                # Temperature color scale (blue to red)
                temp_rule = ColorScaleRule(
                    start_type='min', start_color='87CEEB',
                    mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                    end_type='max', end_color='FF6B6B'
                )
                ws.conditional_formatting.add(f'{temp_col_letter}2:{temp_col_letter}{len(df)+1}', temp_rule)
        
        elif sheet_type == "pond":
            # Water level conditional formatting
            if 'level_cm' in df.columns:
                level_col_idx = df.columns.get_loc('level_cm') + 1
                level_col_letter = openpyxl.utils.get_column_letter(level_col_idx)
                
                # Level color scale
                level_rule = ColorScaleRule(
                    start_type='min', start_color='FFE4B5',
                    end_type='max', end_color='4169E1'
                )
                ws.conditional_formatting.add(f'{level_col_letter}2:{level_col_letter}{len(df)+1}', level_rule)
    
    def _add_excel_charts(self, workbook: openpyxl.Workbook, 
                         data: Dict[str, List[Dict[str, Any]]], 
                         config: AdvancedExportConfig):
        """Add charts to Excel workbook"""
        try:
            # Create charts sheet
            charts_ws = workbook.create_sheet("Charts")
            
            # Add pond level chart if data exists
            if 'pond_metrics' in data and data['pond_metrics']:
                self._add_pond_level_chart(charts_ws, workbook)
            
            # Add temperature chart if data exists
            if 'station_metrics' in data and data['station_metrics']:
                self._add_temperature_chart(charts_ws, workbook)
                
        except Exception as e:
            logger.warning(f"Failed to add Excel charts: {e}")
    
    def _add_pond_level_chart(self, ws: openpyxl.worksheet.worksheet.Worksheet, 
                            workbook: openpyxl.Workbook):
        """Add pond level chart to worksheet"""
        try:
            # Get pond metrics sheet
            pond_ws = workbook["Pond Metrics"]
            
            # Create line chart
            chart = LineChart()
            chart.title = "Pond Water Level Over Time"
            chart.style = 13
            chart.x_axis.title = "Time"
            chart.y_axis.title = "Level (cm)"
            
            # Find level_cm column
            level_col = None
            for idx, cell in enumerate(pond_ws[1], 1):
                if cell.value == 'level_cm':
                    level_col = idx
                    break
            
            if level_col:
                level_col_letter = openpyxl.utils.get_column_letter(level_col)
                data_range = Reference(pond_ws, 
                                     min_col=level_col, max_col=level_col,
                                     min_row=2, max_row=pond_ws.max_row)
                chart.add_data(data_range, titles_from_data=False)
                
                # Add to charts sheet
                ws.add_chart(chart, "A1")
                
        except Exception as e:
            logger.warning(f"Failed to add pond level chart: {e}")
    
    def _add_temperature_chart(self, ws: openpyxl.worksheet.worksheet.Worksheet, 
                             workbook: openpyxl.Workbook):
        """Add temperature chart to worksheet"""
        try:
            # Get station metrics sheet
            station_ws = workbook["Station Metrics"]
            
            # Create line chart
            chart = LineChart()
            chart.title = "Temperature Over Time"
            chart.style = 13
            chart.x_axis.title = "Time"
            chart.y_axis.title = "Temperature (°C)"
            
            # Find temperature_c column
            temp_col = None
            for idx, cell in enumerate(station_ws[1], 1):
                if cell.value == 'temperature_c':
                    temp_col = idx
                    break
            
            if temp_col:
                temp_col_letter = openpyxl.utils.get_column_letter(temp_col)
                data_range = Reference(station_ws,
                                     min_col=temp_col, max_col=temp_col,
                                     min_row=2, max_row=station_ws.max_row)
                chart.add_data(data_range, titles_from_data=False)
                
                # Add to charts sheet (position below pond chart)
                ws.add_chart(chart, "A20")
                
        except Exception as e:
            logger.warning(f"Failed to add temperature chart: {e}")
    
    def _create_analysis_sheet(self, workbook: openpyxl.Workbook, 
                             data: Dict[str, List[Dict[str, Any]]], 
                             config: AdvancedExportConfig):
        """Create data analysis sheet with summary statistics"""
        ws = workbook.create_sheet("Analysis")
        
        # Title
        ws['A1'] = "Data Analysis Summary"
        ws['A1'].font = Font(size=16, bold=True)
        
        row = 3
        
        # Analyze each data type
        for data_type, records in data.items():
            if not records:
                continue
            
            ws[f'A{row}'] = f"{data_type.replace('_', ' ').title()} Analysis"
            ws[f'A{row}'].font = Font(size=12, bold=True)
            row += 1
            
            # Convert to DataFrame for analysis
            df = pd.DataFrame(records)
            
            # Basic statistics
            for col in df.columns:
                if df[col].dtype in ['int64', 'float64']:
                    stats = df[col].describe()
                    
                    ws[f'A{row}'] = f"{col}:"
                    ws[f'A{row}'].font = Font(bold=True)
                    row += 1
                    
                    for stat_name, stat_value in stats.items():
                        ws[f'B{row}'] = f"{stat_name}:"
                        ws[f'C{row}'] = round(stat_value, 2) if not pd.isna(stat_value) else "N/A"
                        row += 1
                    
                    row += 1  # Add spacing
            
            row += 2  # Add more spacing between data types
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    def _aggregate_data(self, data: Dict[str, List[Dict[str, Any]]], 
                       interval: str) -> Dict[str, List[Dict[str, Any]]]:
        """Aggregate data by time interval using pandas"""
        if not pd:
            logger.warning("Pandas not available, skipping aggregation")
            return data
        
        aggregated = {}
        
        for data_type, records in data.items():
            if not records:
                aggregated[data_type] = records
                continue
            
            df = pd.DataFrame(records)
            df['timestamp'] = pd.to_datetime(df['timestamp'])
            df.set_index('timestamp', inplace=True)
            
            # Define aggregation rules
            agg_rules = {}
            for col in df.columns:
                if col in ['level_cm', 'outflow_lps', 'temperature_c', 'battery_v', 'solar_v', 'signal_dbm']:
                    agg_rules[col] = 'mean'
                else:
                    agg_rules[col] = 'first'
            
            # Resample
            if interval == 'hourly':
                resampled = df.resample('H').agg(agg_rules)
            elif interval == 'daily':
                resampled = df.resample('D').agg(agg_rules)
            else:
                resampled = df
            
            # Convert back
            resampled.reset_index(inplace=True)
            aggregated[data_type] = resampled.to_dict('records')
        
        return aggregated
    
    def _update_progress(self, export_id: str, progress: int, message: str, 
                        callback: Callable):
        """Update export progress"""
        try:
            self.export_progress[export_id] = {
                'progress': progress,
                'message': message,
                'timestamp': datetime.now().isoformat()
            }
            
            if callback:
                callback(export_id, progress, message)
                
        except Exception as e:
            logger.warning(f"Failed to update progress: {e}")
    
    def get_export_progress(self, export_id: str) -> Dict[str, Any]:
        """Get export progress status"""
        return self.export_progress.get(export_id, {
            'progress': 0,
            'message': 'Export not found',
            'timestamp': datetime.now().isoformat()
        })


def create_advanced_export_service(db_service: DatabaseService) -> AdvancedExportService:
    """Factory function for advanced export service"""
    return AdvancedExportService(db_service)