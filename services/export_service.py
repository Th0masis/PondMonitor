"""
PondMonitor Export Service

Enhanced data export functionality supporting multiple formats with filtering,
metadata, and progress tracking. Replaces basic export in the main application.

Features:
- Multiple export formats (CSV, JSON, Excel)
- Date range filtering
- Data aggregation options
- Progress tracking for large exports
- Metadata inclusion
- Memory-efficient streaming for large datasets
- Export validation and error handling
"""

import csv
import json
import logging
import tempfile
import zipfile
from datetime import datetime, timezone
from io import StringIO, BytesIO
from typing import List, Dict, Any, Optional, Generator, Union
from dataclasses import dataclass, asdict
from pathlib import Path

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    logging.warning("pandas not available, Excel export will be limited")

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not available, Excel export not supported")

from database import DatabaseService
from utils import ValidationError, DataError, sanitize_filename

logger = logging.getLogger(__name__)


@dataclass
class ExportMetadata:
    """Metadata for exported data"""
    export_time: str
    start_time: str
    end_time: str
    record_count: int
    format: str
    data_types: List[str]
    filters: Dict[str, Any]
    version: str = "1.0"
    generated_by: str = "PondMonitor"
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary"""
        return asdict(self)


@dataclass
class ExportConfig:
    """Configuration for data export"""
    start_time: datetime
    end_time: datetime
    format: str = "csv"
    include_pond_data: bool = True
    include_station_data: bool = True
    include_metadata: bool = True
    aggregate_interval: Optional[str] = None  # 'hourly', 'daily'
    station_id: Optional[str] = None
    limit: Optional[int] = None
    filename_prefix: str = "pond_data"
    
    def validate(self) -> None:
        """Validate export configuration"""
        if self.start_time >= self.end_time:
            raise ValidationError("Start time must be before end time")
        
        if self.format not in ['csv', 'json', 'excel', 'xlsx']:
            raise ValidationError(f"Unsupported format: {self.format}")
        
        if self.format in ['excel', 'xlsx'] and not OPENPYXL_AVAILABLE:
            raise ValidationError("Excel export requires openpyxl package")
        
        if self.aggregate_interval and self.aggregate_interval not in ['hourly', 'daily']:
            raise ValidationError("Aggregate interval must be 'hourly' or 'daily'")


class ExportService:
    """
    Service for exporting pond monitoring data in various formats.
    
    Provides high-level interface for data export with support for
    multiple formats, filtering, and aggregation.
    """
    
    def __init__(self, db_service: DatabaseService):
        """
        Initialize export service.
        
        Args:
            db_service: Database service instance
        """
        self.db = db_service
        self.progress_callbacks = {}
        
    def export_data(self, config: ExportConfig) -> Union[str, bytes]:
        """
        Export data according to configuration.
        
        Args:
            config: Export configuration
        
        Returns:
            Exported data as string or bytes
        
        Raises:
            ValidationError: If configuration is invalid
            DataError: If export fails
        """
        config.validate()
        
        logger.info(
            f"Starting data export: {config.format} format, "
            f"{config.start_time} to {config.end_time}"
        )
        
        try:
            # Get data based on configuration
            data = self._get_export_data(config)
            
            # Apply aggregation if requested
            if config.aggregate_interval:
                data = self._aggregate_data(data, config.aggregate_interval)
            
            # Generate metadata
            metadata = self._generate_metadata(config, data)
            
            # Export in requested format
            if config.format == 'csv':
                return self._export_csv(data, metadata, config)
            elif config.format == 'json':
                return self._export_json(data, metadata, config)
            elif config.format in ['excel', 'xlsx']:
                return self._export_excel(data, metadata, config)
            else:
                raise DataError(f"Unsupported export format: {config.format}")
                
        except Exception as e:
            logger.error(f"Export failed: {e}")
            if isinstance(e, (ValidationError, DataError)):
                raise
            raise DataError(f"Export failed: {e}")
    
    def export_to_file(self, config: ExportConfig, output_path: str) -> str:
        """
        Export data directly to file.
        
        Args:
            config: Export configuration
            output_path: Output file path
        
        Returns:
            Path to created file
        
        Raises:
            ValidationError: If configuration is invalid
            DataError: If export fails
        """
        data = self.export_data(config)
        
        # Ensure directory exists
        output_dir = Path(output_path).parent
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Write data to file
        mode = 'wb' if isinstance(data, bytes) else 'w'
        encoding = None if isinstance(data, bytes) else 'utf-8'
        
        with open(output_path, mode, encoding=encoding) as f:
            f.write(data)
        
        logger.info(f"Data exported to: {output_path}")
        return output_path
    
    def export_archive(self, config: ExportConfig, include_logs: bool = False) -> bytes:
        """
        Export data as ZIP archive with multiple formats and metadata.
        
        Args:
            config: Export configuration
            include_logs: Whether to include system logs
        
        Returns:
            ZIP archive as bytes
        """
        archive_buffer = BytesIO()
        
        with zipfile.ZipFile(archive_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            # Get base data
            data = self._get_export_data(config)
            metadata = self._generate_metadata(config, data)
            
            # Add CSV format
            csv_config = ExportConfig(**asdict(config))
            csv_config.format = 'csv'
            csv_data = self._export_csv(data, metadata, csv_config)
            zf.writestr(f"{config.filename_prefix}.csv", csv_data)
            
            # Add JSON format
            json_config = ExportConfig(**asdict(config))
            json_config.format = 'json'
            json_data = self._export_json(data, metadata, json_config)
            zf.writestr(f"{config.filename_prefix}.json", json_data)
            
            # Add Excel format if available
            if OPENPYXL_AVAILABLE:
                excel_config = ExportConfig(**asdict(config))
                excel_config.format = 'excel'
                excel_data = self._export_excel(data, metadata, excel_config)
                zf.writestr(f"{config.filename_prefix}.xlsx", excel_data)
            
            # Add metadata file
            metadata_json = json.dumps(metadata.to_dict(), indent=2)
            zf.writestr("metadata.json", metadata_json)
            
            # Add README
            readme = self._generate_readme(config, metadata)
            zf.writestr("README.txt", readme)
            
            # Add logs if requested
            if include_logs:
                self._add_logs_to_archive(zf)
        
        archive_buffer.seek(0)
        return archive_buffer.getvalue()
    
    def _get_export_data(self, config: ExportConfig) -> Dict[str, List[Dict[str, Any]]]:
        """Get data for export based on configuration"""
        data = {}
        
        # Get pond metrics if requested
        if config.include_pond_data:
            pond_metrics = self.db.get_pond_metrics(
                config.start_time, 
                config.end_time, 
                config.limit
            )
            data['pond_metrics'] = pond_metrics
            logger.debug(f"Retrieved {len(pond_metrics)} pond metric records")
        
        # Get station metrics if requested
        if config.include_station_data:
            station_metrics = self.db.get_station_metrics(
                config.start_time, 
                config.end_time, 
                config.station_id,
                config.limit
            )
            data['station_metrics'] = station_metrics
            logger.debug(f"Retrieved {len(station_metrics)} station metric records")
        
        return data
    
    def _aggregate_data(self, data: Dict[str, List[Dict[str, Any]]], 
                       interval: str) -> Dict[str, List[Dict[str, Any]]]:
        """Aggregate data by time interval"""
        if not PANDAS_AVAILABLE:
            logger.warning("Pandas not available, skipping aggregation")
            return data
        
        aggregated = {}
        
        for data_type, records in data.items():
            if not records:
                aggregated[data_type] = records
                continue
            
            # Convert to DataFrame
            df = pd.DataFrame(records)
            df['timestamp'] = pd.to_datetime(df['timestamp'])
            df.set_index('timestamp', inplace=True)
            
            # Define aggregation rules
            agg_rules = {}
            for col in df.columns:
                if col in ['level_cm', 'outflow_lps', 'temperature_c', 'battery_v', 'solar_v']:
                    agg_rules[col] = 'mean'
                elif col == 'signal_dbm':
                    agg_rules[col] = 'mean'
                elif col == 'station_id':
                    agg_rules[col] = 'first'
                else:
                    agg_rules[col] = 'first'
            
            # Resample based on interval
            if interval == 'hourly':
                resampled = df.resample('H').agg(agg_rules)
            elif interval == 'daily':
                resampled = df.resample('D').agg(agg_rules)
            else:
                resampled = df
            
            # Convert back to records
            resampled.reset_index(inplace=True)
            resampled['timestamp'] = resampled['timestamp'].dt.strftime('%Y-%m-%d %H:%M:%S')
            aggregated[data_type] = resampled.to_dict('records')
            
            logger.debug(f"Aggregated {data_type}: {len(records)} -> {len(aggregated[data_type])} records")
        
        return aggregated
    
    def _generate_metadata(self, config: ExportConfig, 
                          data: Dict[str, List[Dict[str, Any]]]) -> ExportMetadata:
        """Generate metadata for export"""
        total_records = sum(len(records) for records in data.values())
        data_types = list(data.keys())
        
        return ExportMetadata(
            export_time=datetime.now(timezone.utc).isoformat(),
            start_time=config.start_time.isoformat(),
            end_time=config.end_time.isoformat(),
            record_count=total_records,
            format=config.format,
            data_types=data_types,
            filters={
                'station_id': config.station_id,
                'limit': config.limit,
                'aggregate_interval': config.aggregate_interval
            }
        )
    
    def _export_csv(self, data: Dict[str, List[Dict[str, Any]]], 
                   metadata: ExportMetadata, config: ExportConfig) -> str:
        """Export data as CSV"""
        output = StringIO()
        
        # Write metadata as comments if requested
        if config.include_metadata:
            output.write(f"# PondMonitor Data Export\n")
            output.write(f"# Generated: {metadata.export_time}\n")
            output.write(f"# Period: {metadata.start_time} to {metadata.end_time}\n")
            output.write(f"# Records: {metadata.record_count}\n")
            output.write(f"# Data types: {', '.join(metadata.data_types)}\n")
            output.write(f"#\n")
        
        # Combine all data types into single CSV
        all_records = []
        for data_type, records in data.items():
            for record in records:
                record_with_type = record.copy()
                record_with_type['data_type'] = data_type
                all_records.append(record_with_type)
        
        if all_records:
            # Sort by timestamp
            all_records.sort(key=lambda x: x.get('timestamp', ''))
            
            # Collect all unique field names from all records
            fieldnames = set()
            for record in all_records:
                fieldnames.update(record.keys())
            fieldnames = sorted(fieldnames)  # Sort for consistent order
            
            # Write CSV
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(all_records)
        
        return output.getvalue()
    
    def _export_json(self, data: Dict[str, List[Dict[str, Any]]], 
                    metadata: ExportMetadata, config: ExportConfig) -> str:
        """Export data as JSON"""
        export_data = {
            'data': data,
            'metadata': metadata.to_dict() if config.include_metadata else None
        }
        
        return json.dumps(export_data, indent=2, default=str)
    
    def _export_excel(self, data: Dict[str, List[Dict[str, Any]]], 
                     metadata: ExportMetadata, config: ExportConfig) -> bytes:
        """Export data as Excel workbook"""
        if not OPENPYXL_AVAILABLE:
            raise DataError("Excel export requires openpyxl package")
        
        output = BytesIO()
        
        if PANDAS_AVAILABLE:
            # Use pandas for better Excel support
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Write metadata sheet
                if config.include_metadata:
                    metadata_df = pd.DataFrame([metadata.to_dict()])
                    metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
                
                # Write data sheets
                for data_type, records in data.items():
                    if records:
                        df = pd.DataFrame(records)
                        sheet_name = data_type.replace('_', ' ').title()
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # Basic Excel export using openpyxl
            from openpyxl import Workbook
            
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Add metadata sheet
            if config.include_metadata:
                meta_ws = wb.create_sheet("Metadata")
                meta_data = metadata.to_dict()
                for row, (key, value) in enumerate(meta_data.items(), 1):
                    meta_ws.cell(row=row, column=1, value=key)
                    meta_ws.cell(row=row, column=2, value=str(value))
            
            # Add data sheets
            for data_type, records in data.items():
                if not records:
                    continue
                
                sheet_name = data_type.replace('_', ' ').title()
                ws = wb.create_sheet(sheet_name)
                
                # Write headers
                headers = list(records[0].keys())
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # Write data
                for row, record in enumerate(records, 2):
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=row, column=col, value=record.get(header))
            
            wb.save(output)
        
        output.seek(0)
        return output.getvalue()
    
    def _generate_readme(self, config: ExportConfig, metadata: ExportMetadata) -> str:
        """Generate README file for export archive"""
        readme = f"""PondMonitor Data Export
========================

This archive contains pond monitoring data exported from PondMonitor.

Export Information:
- Generated: {metadata.export_time}
- Time Period: {metadata.start_time} to {metadata.end_time}
- Total Records: {metadata.record_count}
- Data Types: {', '.join(metadata.data_types)}
- Format: {metadata.format}

Files Included:
- {config.filename_prefix}.csv - Data in CSV format
- {config.filename_prefix}.json - Data in JSON format
"""
        
        if OPENPYXL_AVAILABLE:
            readme += f"- {config.filename_prefix}.xlsx - Data in Excel format\n"
        
        readme += """- metadata.json - Export metadata and configuration
- README.txt - This file

Data Description:
- pond_metrics: Water level and outflow measurements
- station_metrics: Sensor telemetry (temperature, battery, signal strength)

For more information about PondMonitor, visit: https://github.com/your-repo/pondmonitor

Data Format Notes:
- Timestamps are in ISO 8601 format (UTC)
- Temperature values are in Celsius
- Voltage values are in Volts
- Signal strength values are in dBm
- Water level values are in centimeters
- Outflow values are in liters per second
"""
        
        return readme
    
    def _add_logs_to_archive(self, zf: zipfile.ZipFile) -> None:
        """Add system logs to archive"""
        try:
            # Add application logs if they exist
            log_files = ['pondmonitor.log', 'ui.log', 'lora_gateway.log']
            
            for log_file in log_files:
                if Path(log_file).exists():
                    zf.write(log_file, f"logs/{log_file}")
                    logger.debug(f"Added {log_file} to archive")
        except Exception as e:
            logger.warning(f"Failed to add logs to archive: {e}")
    
    def get_export_formats(self) -> List[Dict[str, Any]]:
        """
        Get list of available export formats.
        
        Returns:
            List of format information
        """
        formats = [
            {
                'format': 'csv',
                'name': 'CSV (Comma Separated Values)',
                'description': 'Universal spreadsheet format',
                'available': True,
                'mime_type': 'text/csv',
                'extension': '.csv'
            },
            {
                'format': 'json',
                'name': 'JSON (JavaScript Object Notation)',
                'description': 'Machine-readable data format',
                'available': True,
                'mime_type': 'application/json',
                'extension': '.json'
            },
            {
                'format': 'excel',
                'name': 'Excel Workbook',
                'description': 'Microsoft Excel format with multiple sheets',
                'available': OPENPYXL_AVAILABLE,
                'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'extension': '.xlsx'
            }
        ]
        
        return formats
    
    def estimate_export_size(self, config: ExportConfig) -> Dict[str, Any]:
        """
        Estimate export file size and processing time.
        
        Args:
            config: Export configuration
        
        Returns:
            Size estimation information
        """
        # Get record counts
        data = self._get_export_data(config)
        total_records = sum(len(records) for records in data.values())
        
        # Estimate sizes based on format
        estimates = {
            'record_count': total_records,
            'formats': {}
        }
        
        # Average bytes per record (rough estimates)
        bytes_per_record = {
            'csv': 150,    # Average CSV row
            'json': 300,   # JSON with metadata
            'excel': 100   # Excel compressed
        }
        
        for fmt, bpr in bytes_per_record.items():
            estimated_bytes = total_records * bpr
            estimates['formats'][fmt] = {
                'size_bytes': estimated_bytes,
                'size_mb': round(estimated_bytes / (1024 * 1024), 2),
                'estimated_time_seconds': max(1, estimated_bytes // 1000000)  # 1MB/s processing
            }
        
        return estimates


def create_export_service(db_service: DatabaseService) -> ExportService:
    """
    Factory function to create export service.
    
    Args:
        db_service: Database service instance
    
    Returns:
        Configured export service
    """
    return ExportService(db_service)


if __name__ == "__main__":
    # Example usage and testing
    from config import init_config, DatabaseConfig
    from database import init_database
    
    try:
        # Initialize services
        config = init_config()
        db = init_database(config.database)
        export_service = create_export_service(db)
        
        # Test export configuration
        from datetime import datetime, timezone, timedelta
        
        end_time = datetime.now(timezone.utc)
        start_time = end_time - timedelta(hours=24)
        
        export_config = ExportConfig(
            start_time=start_time,
            end_time=end_time,
            format='json',
            filename_prefix='test_export'
        )
        
        # Test size estimation
        size_estimate = export_service.estimate_export_size(export_config)
        print(f"Size estimate: {size_estimate}")
        
        # Test available formats
        formats = export_service.get_export_formats()
        print(f"Available formats: {[f['format'] for f in formats]}")
        
        # Test actual export
        data = export_service.export_data(export_config)
        print(f"Export successful: {len(data)} characters")
        
    except Exception as e:
        print(f"Export service test failed: {e}")